"""Functions for loading and processing project data from Excel."""
import openpyxl
import os
from typing import Dict, List, Tuple, Union, Optional


def load_project_data(filename: str) -> Dict:
    """
    Load project data from Excel file.
    
    Args:
        filename: Path to Excel file
    
    Returns:
        Dictionary containing project data
    """
    wb = openpyxl.load_workbook(filename)
    data_sheet = wb['Data Entry']
    
    # Read PV and EV columns
    pv_series = []
    ev_series = []
    
    for row in data_sheet.iter_rows(min_row=4, min_col=2, max_col=3, values_only=True):
        pv, ev = row
        if pv is None and ev is None:
            break  # stop at end of data
        pv_series.append(pv)
        ev_series.append(ev)
    
    # Get planned duration and start date
    planned_duration = data_sheet['E16'].value  
    start_date = data_sheet['E4'].value  
    
    # Process special markers
    pv_processed = []
    ev_processed = []
    downtime_periods = []  # Periods where PV="XX" (planned downtime)
    stopwork_periods = []  # Periods where EV="XX" (work stoppage)
    
    last_ev = 0.0
    
    for i, (pv, ev) in enumerate(zip(pv_series, ev_series)):
        # Handle PV
        if pv == "XX":
            pv_value = 0.0  # No planned work
            downtime_periods.append(i)
        else:
            pv_value = float(pv)
        
        # Handle EV
        if ev == "XX":
            ev_value = last_ev  # No progress, use last EV
            stopwork_periods.append(i)
        else:
            ev_value = float(ev)
            last_ev = ev_value
        
        pv_processed.append(pv_value)
        ev_processed.append(ev_value)
    
    # Load path data if available
    paths = {}
    try:
        paths_sheet = wb['Paths']
        # Read path definitions - format depends on your actual data structure
        for row in paths_sheet.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                continue
            path_name = row[0]
            path_tasks = row[1].split('-') if row[1] else []
            paths[path_name] = path_tasks
    except:
        # No paths sheet or other issue - for demo we'll create some sample paths
        paths = {
            "Path1-4-8-10": ["1", "4", "8", "10"],
            "Path2-4-8-10": ["2", "4", "8", "10"],
            "Path2-5-9": ["2", "5", "9"],
            "Path3-8-10": ["3", "8", "10"]
        }
    
    # Try to load path-specific PV/EV if available
    path_data = {}
    for path_name in paths:
        try:
            path_sheet = wb[path_name]
            path_pv = []
            path_ev = []
            for row in path_sheet.iter_rows(min_row=4, min_col=2, max_col=3, values_only=True):
                if row[0] is None and row[1] is None:
                    break
                path_pv.append(row[0])
                path_ev.append(row[1])
            
            # Process special markers for path data
            path_pv_processed = []
            path_ev_processed = []
            last_path_ev = 0.0
            
            for pv, ev in zip(path_pv, path_ev):
                # Handle PV
                if pv == "XX":
                    pv_value = 0.0
                else:
                    pv_value = float(pv)
                
                # Handle EV
                if ev == "XX":
                    ev_value = last_path_ev
                else:
                    ev_value = float(ev)
                    last_path_ev = ev_value
                
                path_pv_processed.append(pv_value)
                path_ev_processed.append(ev_value)
                
            path_data[path_name] = {
                'pv': path_pv_processed,
                'ev': path_ev_processed,
                'raw_pv': path_pv,
                'raw_ev': path_ev
            }
        except:
            # No sheet for this path, we'll simulate it later
            pass
    
    return {
        'pv_series': pv_processed,
        'ev_series': ev_processed,
        'raw_pv': pv_series,
        'raw_ev': ev_series,
        'downtime': downtime_periods,
        'stopwork': stopwork_periods,
        'planned_duration': planned_duration,
        'start_date': start_date,
        'paths': paths,
        'path_data': path_data
    }


def simulate_path_data(project_data: Dict) -> Dict:
    """
    Simulate path-specific PV/EV data if not available directly.
    This is a simplified approach for the demo.
    
    Args:
        project_data: Dictionary of loaded project data
    
    Returns:
        Updated project data with simulated path data
    """
    if not project_data['path_data']:
        project_data['path_data'] = {}
    
    pv_series = project_data['pv_series']
    ev_series = project_data['ev_series']
    paths = project_data['paths']
    
    # For each path without data, create simulated data
    for path_name, tasks in paths.items():
        if path_name not in project_data['path_data']:
            # Simple simulation: scale the project PV/EV by a factor based on the path
            # In reality, you would sum PV/EV for all tasks in the path
            scale_factor = len(tasks) / 10.0  # Arbitrary scaling
            
            # Add some variation to make paths differ
            variation = 0.8 + (hash(path_name) % 5) / 10.0  # Between 0.8 and 1.2
            
            path_pv = [pv * scale_factor * variation for pv in pv_series]
            
            # Simulate different performance for different paths
            if "2-4-8-10" in path_name:
                # This path has a work stoppage around period 2
                path_ev = path_pv.copy()
                if len(path_ev) > 2:
                    path_ev[2] *= 0.3  # Major slowdown in period 2
            elif "3-8-10" in path_name:
                # This path starts slow but improves
                path_ev = []
                for i, pv in enumerate(path_pv):
                    if i < 2:
                        path_ev.append(pv * 0.7)  # Slow start
                    else:
                        path_ev.append(pv * 0.9)  # Improves later
            else:
                # Other paths have normal variation
                path_ev = [pv * (0.85 + (i * 0.02)) for i, pv in enumerate(path_pv)]
            
            project_data['path_data'][path_name] = {
                'pv': path_pv,
                'ev': path_ev,
                'raw_pv': path_pv.copy(),
                'raw_ev': path_ev.copy()
            }
    
    return project_data


def write_results_to_excel(project_data: Dict, results: Dict, output_file: str) -> None:
    """
    Write analysis results back to Excel.
    
    Args:
        project_data: Dictionary of project data
        results: Dictionary of analysis results
        output_file: Path to output Excel file
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ES Analysis Results"
    
    # Write headers
    headers = ["Period", "Overall ES", "Overall SPI(t)", "Overall IEAC(t)"]
    for path in results['path_metrics']:
        headers.extend([f"{path} ES", f"{path} SPI(t)", f"{path} IEAC(t)"])
    headers.append("Controlling Path")
    
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    
    # Write data
    for period in range(len(results['overall_metrics'])):
        row_data = [period + 1]  # 1-indexed period number
        
        # Overall metrics
        row_data.extend(results['overall_metrics'][period])
        
        # Path metrics
        for path in results['path_metrics']:
            if period < len(results['path_metrics'][path]):
                row_data.extend(results['path_metrics'][path][period][:-1])  # Exclude SV(t)
            else:
                row_data.extend(["N/A", "N/A", "N/A"])
        
        # Controlling path
        if period < len(results['controlling_path']):
            row_data.append(results['controlling_path'][period])
        else:
            row_data.append("N/A")
        
        # Write row
        for col, value in enumerate(row_data, start=1):
            ws.cell(row=period+2, column=col, value=value)
    
    # Add anomalies if identified
    if 'anomalies' in results:
        ws.cell(row=1, column=len(headers)+1, value="Anomalies")
        for period, anomaly in results['anomalies'].items():
            ws.cell(row=int(period)+2, column=len(headers)+1, 
                    value=f"{anomaly[0]} (IEAC={anomaly[1]:.2f})")
    
    wb.save(output_file)
