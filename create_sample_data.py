"""Create a sample Excel file for ES analysis demonstration"""
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import random

# Create a new workbook and select active worksheet
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Entry"

# Add title and headers
ws['A1'] = "ES Calculator - Sample Project Data"
ws['A1'].font = Font(bold=True, size=14)
ws.merge_cells('A1:F1')
ws['A1'].alignment = Alignment(horizontal='center')

ws['A3'] = "Period"
ws['B3'] = "Cumulative PV"
ws['C3'] = "Cumulative EV"
ws['A3'].font = Font(bold=True)
ws['B3'].font = Font(bold=True)
ws['C3'].font = Font(bold=True)

# Add project information
ws['D4'] = "Start Date:"
ws['E4'] = datetime(2025, 1, 1)  # January 1, 2025
ws['D16'] = "Planned Duration (PD):"
ws['E16'] = 10  # 10 weeks

# Generate sample PV data (cumulative)
pv_data = [0]  # Start with 0
for i in range(10):
    if i == 5:  # Add a planned downtime at period 6
        pv_data.append("XX")
    else:
        inc = random.uniform(8, 12)  # Random increment between 8-12
        last_val = pv_data[-1] if pv_data[-1] != "XX" else pv_data[-2]
        if isinstance(last_val, str):
            last_val = 0
        pv_data.append(last_val + inc)

# Generate sample EV data (cumulative, typically below PV)
ev_data = [0]  # Start with 0
for i in range(10):
    if i == 2:  # Add a work stoppage at period 3
        ev_data.append("XX")
    else:
        if pv_data[i+1] == "XX":  # If PV is downtime, EV is also downtime
            ev_data.append("XX")
        else:
            perf_factor = random.uniform(0.85, 0.98)  # Performance factor
            last_val = ev_data[-1] if ev_data[-1] != "XX" else ev_data[-2]
            if isinstance(last_val, str):
                last_val = 0
            planned = pv_data[i+1] if pv_data[i+1] != "XX" else pv_data[i]
            if isinstance(planned, str):
                planned = 0
            ev_data.append(last_val + (planned - last_val) * perf_factor)

# Write data to worksheet
for i in range(11):  # 0-10 periods
    ws[f'A{i+4}'] = i
    
    # PV data
    if i < len(pv_data):
        ws[f'B{i+4}'] = pv_data[i]
    
    # EV data
    if i < len(ev_data):
        ws[f'C{i+4}'] = ev_data[i]

# Create Paths sheet with path definitions
paths_sheet = wb.create_sheet("Paths")
paths_sheet['A1'] = "Path Name"
paths_sheet['B1'] = "Tasks"
paths_sheet['A1'].font = Font(bold=True)
paths_sheet['B1'].font = Font(bold=True)

# Define some sample paths
paths = [
    ("Path1-4-8-10", "1-4-8-10"),
    ("Path2-4-8-10", "2-4-8-10"),
    ("Path2-5-9", "2-5-9"),
    ("Path3-8-10", "3-8-10")
]

for i, (path_name, tasks) in enumerate(paths):
    paths_sheet[f'A{i+2}'] = path_name
    paths_sheet[f'B{i+2}'] = tasks
    
    # Create a sheet for each path
    path_sheet = wb.create_sheet(path_name)
    
    # Add headers
    path_sheet['A1'] = f"Path: {path_name}"
    path_sheet['A1'].font = Font(bold=True, size=12)
    path_sheet.merge_cells('A1:C1')
    path_sheet['A1'].alignment = Alignment(horizontal='center')
    
    path_sheet['A3'] = "Period"
    path_sheet['B3'] = "Cumulative PV"
    path_sheet['C3'] = "Cumulative EV"
    path_sheet['A3'].font = Font(bold=True)
    path_sheet['B3'].font = Font(bold=True)
    path_sheet['C3'].font = Font(bold=True)
    
    # Generate path-specific data based on the main data with variations
    scale = random.uniform(0.3, 0.6)  # Scale factor for this path
    
    # Generate path PV and EV
    path_pv = []
    path_ev = []
    
    for j in range(11):
        path_sheet[f'A{j+4}'] = j
        
        if j < len(pv_data):
            # Handle special case for Path2-4-8-10 (major slowdown in period 2)
            if path_name == "Path2-4-8-10" and j == 2:
                perf_multiplier = 0.3  # Much worse performance
            elif path_name == "Path3-8-10" and j < 2:
                perf_multiplier = 0.7  # Slow start
            elif path_name == "Path3-8-10" and j >= 2:
                perf_multiplier = 0.9  # Improves later
            else:
                perf_multiplier = random.uniform(0.85, 0.98)
            
            # PV for path
            if pv_data[j] == "XX":
                path_sheet[f'B{j+4}'] = "XX"
                path_pv.append("XX")
            else:
                scaled_pv = pv_data[j] * scale
                path_sheet[f'B{j+4}'] = scaled_pv
                path_pv.append(scaled_pv)
            
            # EV for path
            if ev_data[j] == "XX" or pv_data[j] == "XX":
                path_sheet[f'C{j+4}'] = "XX"
                path_ev.append("XX")
            else:
                # For Path2-4-8-10, make period 2 a major slowdown
                if path_name == "Path2-4-8-10" and j == 2:
                    last_ev = path_ev[-1] if path_ev and path_ev[-1] != "XX" else 0
                    if isinstance(last_ev, str):
                        last_ev = 0
                    scaled_ev = last_ev + (path_pv[j] - last_ev) * 0.3  # Major slowdown
                else:                  
                    last_ev = path_ev[-1] if path_ev and path_ev[-1] != "XX" else 0
                    if isinstance(last_ev, str):
                        last_ev = 0
                    curr_pv = path_pv[j] if path_pv[j] != "XX" else path_pv[j-1] if j > 0 else 0
                    if isinstance(curr_pv, str):
                        curr_pv = 0
                    scaled_ev = last_ev + (curr_pv - last_ev) * perf_multiplier
                
                path_sheet[f'C{j+4}'] = scaled_ev
                path_ev.append(scaled_ev)

# Save the workbook
wb.save('ES Calculator Longest Path.xlsx')
print("Sample Excel file 'ES Calculator Longest Path.xlsx' created successfully!")
